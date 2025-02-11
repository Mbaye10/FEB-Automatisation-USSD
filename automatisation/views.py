from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta
from .models import CanalUSSD, UseCase, TestResult, AlertConfiguration, DetailedTestResult
from django.db.models import Count
from django.db.models.functions import TruncDate
import json
from .utils import get_dashboard_stats, get_failed_tests_details, export_to_excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def index(request):
    # Récupérer les statistiques des 7 derniers jours
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    
    # Récupérer les statistiques de base
    total_tests = TestResult.objects.filter(
        date_test__range=[start_date, end_date]
    ).count()
    
    tests_ok = TestResult.objects.filter(
        date_test__range=[start_date, end_date],
        statut='OK'
    ).count()
    
    tests_nok = total_tests - tests_ok
    taux_dysfonctionnement = (tests_nok / total_tests * 100) if total_tests > 0 else 0
    
    # Récupérer les tests en échec récents
    tests_en_echec = TestResult.objects.filter(
        date_test__range=[start_date, end_date],
        statut='NOK'
    ).select_related('use_case', 'use_case__canal').order_by('-date_test')[:10]
    
    # Récupérer les canaux pour le filtre
    canaux = CanalUSSD.objects.all()
    
    context = {
        'total_tests': total_tests,
        'tests_ok': tests_ok,
        'tests_nok': tests_nok,
        'taux_dysfonctionnement': round(taux_dysfonctionnement, 2),
        'tests_en_echec': tests_en_echec,
        'canaux': canaux,
    }
    
    return render(request, 'automatisation/index.html', context)

def login_view(request):
    return render(request, 'login.html')

def home_redirect(request):
    return redirect('index')

@login_required
def dashboard(request):
    # Récupération des paramètres de filtrage
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    canal_id = request.GET.get('canal')
    
    # Si pas de dates spécifiées, utiliser les 7 derniers jours
    if not start_date:
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d') if end_date else datetime.now()
    
    # Récupération des statistiques
    stats = get_dashboard_stats(start_date, end_date, canal_id)
    failed_tests = get_failed_tests_details(start_date, end_date, canal_id)
    
    # Données pour le graphique d'évolution
    evolution_data = TestResult.objects.filter(
        date_test__range=[start_date, end_date]
    ).annotate(
        date=TruncDate('date_test')
    ).values('date', 'statut').annotate(
        count=Count('id')
    ).order_by('date')
    
    # Préparation des données pour les graphiques
    dates = []
    ok_counts = []
    nok_counts = []
    current_date = start_date
    
    while current_date <= end_date:
        dates.append(current_date.strftime('%Y-%m-%d'))
        ok_count = 0
        nok_count = 0
        
        for entry in evolution_data:
            if entry['date'].strftime('%Y-%m-%d') == current_date.strftime('%Y-%m-%d'):
                if entry['statut'] == 'OK':
                    ok_count = entry['count']
                else:
                    nok_count = entry['count']
        
        ok_counts.append(ok_count)
        nok_counts.append(nok_count)
        current_date += timedelta(days=1)
    
    # Données pour le graphique de distribution par canal
    canal_stats = TestResult.objects.filter(
        date_test__range=[start_date, end_date]
    ).values(
        'use_case__canal__code'
    ).annotate(
        count=Count('id')
    )
    
    canal_labels = [stat['use_case__canal__code'] for stat in canal_stats]
    canal_data = [stat['count'] for stat in canal_stats]
    
    # Récupération des canaux pour le filtre
    canaux = CanalUSSD.objects.all()
    use_cases = UseCase.objects.all().select_related('canal')
    
    context = {
        'stats': stats,
        'failed_tests': failed_tests,
        'canaux': canaux,
        'use_cases': use_cases,
        'selected_canal': canal_id,
        'start_date': start_date,
        'end_date': end_date,
        'evolution_dates': json.dumps(dates),
        'evolution_ok': json.dumps(ok_counts),
        'evolution_nok': json.dumps(nok_counts),
        'canal_labels': json.dumps(canal_labels),
        'canal_data': json.dumps(canal_data),
    }
    
    return render(request, 'automatisation/dashboard.html', context)

@login_required
def export_results(request):
    # Récupération des paramètres de filtrage
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    canal_id = request.GET.get('canal')
    
    # Conversion des dates
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d') if end_date else datetime.now()
    
    # Récupération des tests en échec
    failed_tests = get_failed_tests_details(start_date, end_date, canal_id)
    
    # Création du fichier Excel
    df = export_to_excel(failed_tests)
    
    # Génération de la réponse HTTP
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="rapport_tests_ussd.xlsx"'
    
    # Écriture du DataFrame dans le fichier Excel
    df.to_excel(response, index=False)
    
    return response

@login_required
def alert_configuration(request):
    if request.method == 'POST':
        use_case_id = request.POST.get('use_case')
        email = request.POST.get('email')
        
        # Création ou mise à jour de la configuration d'alerte
        AlertConfiguration.objects.update_or_create(
            user=request.user,
            use_case_id=use_case_id,
            email=email,
            defaults={'is_active': True}
        )
    
    # Récupération des configurations existantes
    user_alerts = AlertConfiguration.objects.filter(user=request.user)
    use_cases = UseCase.objects.all()
    
    context = {
        'user_alerts': user_alerts,
        'use_cases': use_cases,
    }
    
    return render(request, 'automatisation/alert_configuration.html', context)

@login_required
def audits_tests(request):
    # Récupération des paramètres de filtrage
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    canal_id = request.GET.get('canal')
    status = request.GET.get('status')

    # Initialisation de la requête
    tests = TestResult.objects.all().select_related('use_case', 'use_case__canal')

    # Application des filtres
    if start_date:
        tests = tests.filter(date_test__gte=start_date)
    if end_date:
        tests = tests.filter(date_test__lte=end_date)
    if canal_id:
        tests = tests.filter(use_case__canal_id=canal_id)
    if status:
        tests = tests.filter(status=status)

    # Récupération des canaux pour le filtre
    canaux = CanalUSSD.objects.all()

    context = {
        'tests': tests,
        'canaux': canaux,
        'start_date': start_date,
        'end_date': end_date,
        'selected_canal': int(canal_id) if canal_id else None,
        'selected_status': status,
    }

    return render(request, 'automatisation/audits_tests.html', context)

@login_required
def export_audit(request):
    # Récupération des paramètres de filtrage
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    canal_id = request.GET.get('canal')
    status = request.GET.get('status')

    # Initialisation de la requête
    tests = TestResult.objects.all().select_related('use_case', 'use_case__canal')

    # Application des filtres
    if start_date:
        tests = tests.filter(date_test__gte=start_date)
    if end_date:
        tests = tests.filter(date_test__lte=end_date)
    if canal_id:
        tests = tests.filter(use_case__canal_id=canal_id)
    if status:
        tests = tests.filter(status=status)

    # Création du fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit des Tests"

    # En-têtes
    headers = ['Date/Heure', 'Canal', 'Use Case', 'Numéro', 'Statut', 'Durée (s)', 'Message']
    ws.append(headers)

    # Données
    for test in tests:
        row = [
            test.date_test.strftime("%d/%m/%Y %H:%M:%S"),
            test.use_case.canal.code,
            test.use_case.nom,
            test.numero_test,
            test.status,
            f"{test.duree_test:.2f}",
            test.message_erreur or "-"
        ]
        ws.append(row)

    # Formatage
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Création de la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=audit_tests.xlsx'

    # Sauvegarde du fichier
    wb.save(response)
    return response